import openpyxl

class checkTemplate():
    def __init__(self, filename):
        self.Settings = {
            'Федеральные услуги': {'RowUslugi': 'E', 'RowOrganisation': 'F', 'Cell': 8, 'Status': True},
            'Иные услуги': {'RowUslugi': 'E', 'RowOrganisation': 'F', 'Cell': 8, 'Status': True},
            'Региональные услуги': {'RowUslugi': 'E', 'RowOrganisation': 'F', 'Cell': 8, 'Status': True},
            'Муниципальные услуги': {'RowUslugi': 'E', 'RowOrganisation': 'F', 'Cell': 8, 'Status': True},
            'Окна': {'RowUslugi': None, 'RowOrganisation': None, 'Cell': None, 'Status': False}
        }                                           # Получаем настройки на каждый тип шаблона
        self.CellSettingsType = {
            'Услуги федеральных органов власти': ['E1', 'Федеральные услуги', 'F3'],
            'Услуги региональных органов власти': ['E1', 'Региональные услуги', 'F3'],
            'Сведения о количестве действующих окон и специалистов в': ['C1', 'Окна', 'D3'],
            'Услуги муниципальных органов власти': ['E1', 'Муниципальные услуги', 'F3'],
            'АО «Корпорация «МСП»': ['F8', 'Иные услуги', 'F3'],
            }                                       # Получаем тип шаблона
        self.name_type = ''                         # Получаем название шаблона
        self.type_orgs = None                       # Получаем тип организаций шаблона
        self.startRowCell = ['A', 'C']              # A - номер АИС МРС, B - Адрес МФЦ, C - тип МФЦ.
        self.CellInfoTypeTemplate = ['C1']          # Получение типа шаблона
        self.startIndex = 1                         # Начальный индекс поиска информации по шаблону.
        self.wb = openpyxl.load_workbook(filename)  # Открываем файл

    def GetTypeTemplate(self):
        try:
            _info = self.wb["2833"]    # Открываем таблицу 2833(потому что есть что в МФЦ)
        except Exception as err:
            _info = self.wb["9004288"]  # Открываем таблицу 9004288(потому что есть что в TOSP)

        for key_type in self.CellSettingsType.keys():
            try:
                settings = self.CellSettingsType[key_type]
                cell = settings[0]
                name_t = settings[1]
                self.type_orgs = True if 'МФЦ' in _info[settings[2]].value else False
                data_cell = _info[cell].value
                if data_cell:
                    if key_type in data_cell:
                        self.name_type = name_t
                        return True
            except Exception as err:
                pass
        return False

    def GetAllMFC(self):
        _info = self.wb["Справочная информация"]    # Находим книгу "Справочная информация"

        all_list_mfc = []
        # Перебираем столбцы по названию и индексу
        for num_cells in range(1, _info.max_row):
            temp_list = {}
            self.startIndex += 1
            # Получаем данные с заданного диапазона
            mfc_data = _info[self.startRowCell[0] + str(self.startIndex) : self.startRowCell[1] + str(self.startIndex)]
            # Перебираем данные пропуская None и записываем с структуированный список JSON
            for val_cell in mfc_data:
                if val_cell[0].value:
                    temp_list['numAisMrs'] = val_cell[0].value
                    temp_list['nameMfc'] = val_cell[1].value
                    temp_list['typeMfc'] = str(val_cell[2].value).lower()
                    all_list_mfc.append(temp_list)
                    temp_list = {}
        return all_list_mfc, len(all_list_mfc)

    def GetAllDataUsl(self):
        _info = self.wb.active  # Выбираем 1-ю книгу
        all_list_usl = []
        all_list_org = []
        # Получаем настройки шаблона ВИД: {'RowUslugi': 'E', 'RowOrganisation': 'F', 'Cell': 8}
        get_settings = self.Settings[self.name_type]
        if get_settings['Status']:
            for num_cells in range(0, _info.max_row):
                temp_list = {}
                usl_data = _info[get_settings['RowUslugi'] + str(get_settings['Cell']): get_settings['RowOrganisation'] + str(get_settings['Cell'])]
                for val_cell in usl_data:
                    if val_cell[0].value:
                        usl = val_cell[0].value
                        org = val_cell[1].value
                        temp_list['nameUsl'] = usl
                        temp_list['organisation'] = org
                        # ----- расширяемо
                        if "Росреестр" in org:
                            temp_list['cell_in_0'] = "G{0}".format(get_settings['Cell'])
                            temp_list['cell_in_exter'] = "H{0}".format(get_settings['Cell'])
                            temp_list['cell_out_exter'] = "M{0}".format(get_settings['Cell'])
                            temp_list['cell_out_0'] = "L{0}".format(get_settings['Cell'])
                            temp_list['cell_comment'] = "T{0}".format(get_settings['Cell'])
                        else:
                            temp_list['cell_in_0'] = "K{0}".format(get_settings['Cell'])
                            temp_list['cell_in_exter'] = ""
                            temp_list['cell_out_exter'] = ""
                            temp_list['cell_out_0'] = "R{0}".format(get_settings['Cell'])
                            temp_list['cell_comment'] = "T{0}".format(get_settings['Cell'])
                        temp_list['cell_all_konsult'] = "S{0}".format(_info.max_row)
                        # Добавляем в список названия организаций
                        if org not in all_list_org:
                            all_list_org.append(org)
                        all_list_usl.append(temp_list)
                get_settings['Cell'] += 1
            return all_list_usl, len(all_list_usl), all_list_org, len(all_list_org)
        return None, 0, None, 0

    def closeTemplate(self):
        self.wb.close()